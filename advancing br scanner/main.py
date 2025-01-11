import cv2
from pyzbar.pyzbar import decode
import openpyxl
from datetime import datetime
from tkinter import Tk, Label, Button, StringVar, OptionMenu

# Initialize or create an Excel file for a subject
def initialize_excel(subject):
    excel_file = f"{subject}.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Barcode", "Date", "Time"])
        workbook.save(excel_file)
    return excel_file

# Update attendance in the Excel file
def mark_attendance(barcode_data, name, subject):
    excel_file = initialize_excel(subject)
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Attendance"]
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")

    # Check if entry already exists for the same date
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == barcode_data and row[2] == current_date:
            return False  # Entry already exists

    # Add new entry
    sheet.append([name, barcode_data, current_date, current_time])
    workbook.save(excel_file)
    return True

# Barcode scanning function
def scan_barcode(subject):
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        for barcode in decode(frame):
            barcode_data = barcode.data.decode('utf-8')
            barcode_type = barcode.type

            # Draw a rectangle around the barcode
            points = barcode.polygon
            if len(points) == 4:
                pts = [(point.x, point.y) for point in points]
                cv2.polylines(frame, [pts], isClosed=True, color=(0, 255, 0), thickness=2)

            # Display barcode data
            cv2.putText(frame, f"{barcode_data} ({barcode_type})", (barcode.rect.left, barcode.rect.top - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

            # Update attendance
            if mark_attendance(barcode_data, "User", subject):  # Replace "User" with actual name lookup if needed
                show_message(f"Attendance marked for {subject}!")
            else:
                show_message(f"Attendance already marked for {subject} today!")

        cv2.imshow("Barcode Scanner", frame)

        # Exit with 'q'
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# GUI to select subject and start scanning
def select_subject():
    def start_scanning():
        subject = subject_var.get()
        if subject:
            root.destroy()
            scan_barcode(subject)
        else:
            show_message("Please select a subject!")

    root = Tk()
    root.title("Select Subject")
    root.geometry("300x200")

    subject_var = StringVar(root)
    subject_var.set("")  # Default value

    Label(root, text="Select Subject:", font=("Arial", 14)).pack(pady=10)
    subjects = ["Math", "Science", "English", "History"]  # Add your subjects here
    OptionMenu(root, subject_var, *subjects).pack(pady=10)

    Button(root, text="Start Scanning", command=start_scanning, font=("Arial", 12), pady=5).pack(pady=20)
    root.mainloop()

# GUI feedback
def show_message(message):
    root = Tk()
    root.title("Attendance Status")
    label = Label(root, text=message, font=("Arial", 14), padx=20, pady=20)
    label.pack()
    Button(root, text="OK", command=root.destroy, font=("Arial", 12), pady=10).pack()
    root.mainloop()

# Main Function
if __name__ == "__main__":
    select_subject()
